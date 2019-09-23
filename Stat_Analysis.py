# -*- coding: utf-8 -*-
"""
Created on Fri Aug  2 13:43:22 2019

@author: cxy
"""

import pandas as pd
from openpyxl import load_workbook
from configparser import ConfigParser
from os import listdir, system
from datetime import datetime
from copy import copy

cfg = ConfigParser()
cfg.read('ACCT.ini',encoding='utf-8-sig')
code_1 = [code.strip() for code in cfg.get('ACCT1','code').split(',')]
code_2 = [code.strip() for code in cfg.get('ACCT2','code').split(',')]
BUY_RATIO = float(cfg.get('HKD','buy_ratio'))
SELL_RATIO = float(cfg.get('HKD','sell_ratio'))
CLOSE_RATIO = float(cfg.get('HKD','close_ratio'))
FLAG = cfg.get('AUTO_CALC','flag')
KEEP = cfg.get('FUTURE_COL','keep')
PATH = cfg.get('DIR','path')
SECS = int(cfg.get('TIME','secs'))
COMBS = [comb.strip() for comb in cfg.get('COMB','comb_list').split(',')]
INDUSTRY = pd.read_excel('INDUSTRY.xlsx',sheet_name=0)
INDUSTRY.iloc[:,0] = INDUSTRY.iloc[:,0].apply(lambda x: x[:6])
INDUSTRY.drop(columns=['证券简称'],inplace=True)
date_str = datetime.today().strftime('%Y%m%d')
files = listdir()
print('正在处理' + '.' * 64)

if FLAG.upper().startswith(('F','0')):
    print('Loading file.')
    FLAG = False
else:
    FLAG = True
    print('Auto Calc.')

if KEEP.upper().startswith(('F','0')):
    KEEP = False
else:
    KEEP = True

for file in files:
    if file.startswith('持仓统计'):
        rd_file = file
    elif file.startswith('组合证券') and not FLAG:
        file_holding = file
    elif file.startswith('期货保证金'):
        file_analys = file
    elif file.startswith('期货持仓'):
        file_futures = file
    elif file.startswith('新综合信息查询'):
        comb_file = file
        

wb = load_workbook(rd_file)
wb_data = load_workbook(rd_file,data_only=True)


'''
wb.sheetnames

ws_1 = wb['持仓明细表']
Holdings = pd.DataFrame(list(ws_1.values)[1:],columns=list(ws_1.values)[0])
Sub_hold = Holdings[['组合名称','证券名称','证券代码','市值','成本']]
Sub_hold.groupby('组合名称')[['市值','成本']].sum()
sql_acct_1 = "组合名称 == {!r} or (组合名称 == {!r} and 证券代码 in {!r})".format('权益类6号组合','股票组合1',code_1)
Holdings.query(sql_acct_1).sort_values(['组合名称','市值'],ascending=False)
sql_acct_2 = "组合名称 == {!r} or (组合名称 == {!r} and 证券代码 in {!r})".format('权益类5号组合','股票组合1',code_2)
Holdings.query(sql_acct_2).sort_values(['组合名称','市值'],ascending=False)
'''
sql_acct_0 = "(组合名称 == {!r} and 证券代码 not in {!r} and 证券代码 not in {!r}) or 组合名称 in {!r}".format('股票组合1',code_1,code_2,['权益类2号组合','权益类3号组合'])
sql_acct_1 = "组合名称 in {!r} or (组合名称 == {!r} and 证券代码 in {!r})".format(['权益类6号组合','沪港通组合'],'股票组合1',code_1)
sql_acct_2 = "组合名称 == {!r} or (组合名称 == {!r} and 证券代码 in {!r})".format('权益类5号组合','股票组合1',code_2)
sql_acct_3 = "组合名称.str.contains({!r})".format('量化')

######### 

#ws_3 = wba['单账户1']
#ws_data = wb_data['单账户1']

#ws_3.max_row

def Get_DF(ws,start_cell,end_col):
    '''Params: ws- Sheet 
    Range- Sheet Range
    Rts: DataFrame
    '''
    tmp = []
    Range = start_cell + ':{}{}'.format(end_col,ws.max_row)
    for row in ws[Range]:
        tmp_in = []
        for col in row:
            if col.value == None:
                break
            tmp_in.append(col.value)
        if not tmp_in:
            break
        tmp.append(tmp_in)
    return pd.DataFrame(tmp[1:],columns=tmp[0])

def Clear_Range(ws,Range):
    for row in ws[Range]:
        for col in row:
            col.value = None

def Write_DF(ws,cell,df,header=False):
    if not header:
        Range = cell + ':{}'.format(ws[cell].offset(*df.shape).offset(-1,-1).coordinate)
        val_row = iter(df.values)
    else:
        val_row = iter([df.columns.tolist()] + df.values.tolist())  
        Range = cell + ':{}'.format(ws[cell].offset(*df.shape).coordinate)
    for row in ws[Range]:
        val_col = iter(next(val_row))
        for col in row:
            try:
                col.value = next(val_col)
            except:
                continue
        

'''
tmp = []
for row in ws_data['A2:K{}'.format(ws_data.max_row)]:
    tmp_in = []
    for col in row:
        if col.value == None:
            break
        tmp_in.append(col.value)
    if not tmp_in:
        break
    tmp.append(tmp_in)
pd.DataFrame(tmp[1:],columns=tmp[0]).columns
'''


tmp = pd.read_excel(comb_file)
#################################################### 
if FLAG:
#cols = ['组合名称', '证券代码', '数量', '成本价','成本', '浮动盈亏','投资收益', '累积结转', '累积投资收益', '买入费用', '卖出费用']
    cost_ws = Get_DF(wb['持仓明细表'],'A1','U')
    out_cols = cost_ws.columns
    pd.options.mode.chained_assignment = None # default='warn'
    #tmp = pd.read_excel(comb_file)
    cols = ['组合名称','证券名称','证券代码','最新价','持仓','当日买量','红股权益','当日买金额','当日卖量','当日卖金额','当日买费用','当日卖费用','费用合计','当日红利','证券类别','当日红股','当前成本','市值']
    
    stocks_trades = tmp[cols].query("(组合名称 in {!r} and 证券类别 in ['股票', '开放式基金']) and (持仓 > 0 or 费用合计 > 0)".format(COMBS))
    stocks_trades.rename(columns={'市值': '市值2'}, inplace=True)
    stocks_trades.loc[stocks_trades.组合名称 == '沪港通组合',['当日买金额','当日买费用']] *= BUY_RATIO
    stocks_trades.loc[stocks_trades.组合名称 == '沪港通组合',['当日卖金额','当日卖费用']] *= SELL_RATIO
    #code_name = dict(zip(stocks_trades['证券代码'],stocks_trades['证券名称']))
    code_name = dict(zip(cost_ws['证券代码'],cost_ws['证券名称']))
    code_updated = dict(zip(tmp['证券代码'],tmp['证券名称']))
    code_name.update(code_updated)
    stocks_trades['当日买量'] += stocks_trades['红股权益'] + stocks_trades['当日红股']
    stocks_trades.drop(columns=['证券名称','红股权益','费用合计','证券类别'],inplace=True)
    #stocks_trades = stocks_trades[['组合名称','证券代码','最新价','当日买量','当日买金额','当日卖量','当日卖金额','当日买费用','当日卖费用','费用合计']]
    #(净买金额 != 0 or 费用合计 != 0) and
    #stocks_trades = stocks_trades.query("持仓 > 0 or 费用合计 > 0")
    cost_merg = pd.merge(cost_ws,stocks_trades,on=['组合名称','证券代码'],how='outer')
    cost_merg.fillna(0,inplace=True)
    #cost_merg.loc[cost_merg.证券名称 == 0,'证券名称'] = cost_merg.loc[cost_merg.证券名称 == 0,'证券代码'].apply(lambda x: code_name.get(x))
    cost_merg['证券名称'] = cost_merg['证券代码'].apply(lambda x: code_name.get(x))
    
    unchg_0 = cost_merg.query("持仓 == 数量 == 0")[out_cols]
    others = cost_merg.query("持仓 >0 or 数量 > 0")
    
    ###成本
    others['成本'] += others['当日买金额']
    
    ####成本价
    others['成本价'] = others['成本'] / (others['数量'] + others['当日买量'])
    
    others['成本'] -= others['成本价'] * others['当日卖量']
    
    ###市值
    others['市值'] = others['持仓'] * others['最新价']
    
    ### 结转利润： 新浮动盈亏 - 浮动盈亏 
    others['结转利润'] = others['市值'] - others['成本'] - others['浮动盈亏']
    
    #### 浮动盈亏
    others['浮动盈亏'] = others['市值']- others['成本']
    
    ####新累积结转 ： 累积结转 +  新的结转利润
    others['累积结转'] += others['结转利润']
    
    #### 投资收益： 投资收益 + 卖出金额 - 卖出成本  
    others['投资收益'] = others['当日卖金额'] - others['成本价'] * others['当日卖量'] 
    
    #### 累积投资收益 : 累积投资收益 + 投资收益
    others['累积投资收益'] += others['投资收益']
    
    #### 总盈亏 ： 累积结转 + 累积投资收益
    others['总盈亏'] = others['累积投资收益'] + others['累积结转']
    
    #####买入费用 : 买入费用 + 新买入费用
    others['买入费用'] += others['当日买费用']
    
    #######卖出费用 : 卖出费用 + 新卖出费用
    others['卖出费用'] += others['当日卖费用']
    
    ######总费用： 买入费用 + 卖出费用
    others['总费用'] = others['买入费用'] + others['卖出费用']
    
    ##### 费后总盈利： 总盈亏 - 总费用
    others['费后总盈利'] = others['总盈亏'] - others['总费用']

    others['数量'] = others['持仓']
    #others['收盘价'] = others['最新价']
    others['分红'] += others['当日红利']
    
    in_sht_cost = others['当前成本'].sum()
    in_sht_mkt_val = others['市值2'].sum()
    print('计算成本：{:.2f}, O3内成本：{:.2f}'.format(others['成本'].sum(), in_sht_cost))
    if int(in_sht_cost) != int(others['成本'].sum()):
        print('\033[31m成本不一致.\033[0m')
    
    print('计算市值：{:.2f}, O3内成本：{:.2f}'.format(others['市值'].sum(), in_sht_mkt_val))
    if others['市值'].sum() != in_sht_mkt_val:
        print('\033[31m市值不一致.\033[0m')
    
    others = others[out_cols]
    sheet_1 = others.append(unchg_0)
    
    sheet_1[out_cols[-1]] = sheet_1['费后总盈利'] + sheet_1['分红']
    
    #sheet_1['收盘价'] = others['最新价']
    close_dict = pd.Series(tmp['最新价'].values,index=tmp['证券代码']).dropna().to_dict()
    #[(key,val) for key, val in close_dict.items() if len(key) == 5]
    [close_dict.update({key: val * CLOSE_RATIO}) for key, val in close_dict.items() if len(key) == 5]
    
    sheet_1['收盘价'] = sheet_1['证券代码'].apply(lambda x: close_dict.get(x))
    sheet_1['统计日期'] = date_str
    
    # fillna for column close
    sheet_1.fillna({'收盘价': 0}, inplace=True)
    
    #n_row = wb['持仓明细表'].max_row
    #Clear_Range(wb['持仓明细表'],'A2:S{}'.format(n_row))
    #Write_DF(wb['持仓明细表'],'A2',sheet_1)
    
    #sheet_1.head()
    #sheet_1.sort_values('费后总盈利')
    #[x for x in sheet_1.groupby('组合名称')[['市值','成本','累积结转','累积投资收益','总盈亏','总费用','费后总盈利']].sum().sum()]
    #[x for x in sheet_1.query("证券代码 in {!r}".format(sheet_1.query("数量 > 0")['证券代码'].values.tolist())).groupby('证券代码')[['市值','成本','累积结转','累积投资收益','总盈亏','总费用','费后总盈利']].sum().sum()]
    sheet_2 = sheet_1.query("not 组合名称.str.contains('量化')").groupby(['证券名称','证券代码'])[['数量','成本','市值','累积结转','累积投资收益','总盈亏','总费用','费后总盈利'] + list(out_cols[-2:])].sum().query("数量 > 0").reset_index()
    #[x for x in sheet_1.query("数量 > 0").groupby('证券代码')[['市值','成本','累积结转','累积投资收益','总盈亏','总费用','费后总盈利']].sum().sum()]
    sheet_2.insert(0,'统计日期',date_str)
    sheet_2.insert(11,'盈亏率',sheet_2[out_cols[-1]] / sheet_2['成本'])
    #n_row = wb['权益类多头汇总明细表'].max_row
    #Clear_Range(wb['权益类多头汇总明细表'],'A2:L{}'.format(n_row))
    #Write_DF(wb['权益类多头汇总明细表'],'A2',sheet_2)
    '''
    结转利润： 新浮动盈亏 - 浮动盈亏 
    新累积结转 ： 累积结转 +  新的结转利润
    投资收益： 投资收益 + 卖出金额 - 卖出成本 
    累积投资收益 : 累积投资收益 + 投资收益
    总盈亏 ： 累积结转 + 累积投资收益
    买入费用 : 买入费用 + 新买入费用
    卖出费用 : 卖出费用 + 新卖出费用
    总费用： 买入费用 + 卖出费用
    费后总盈利： 总盈亏 - 总费用
    '''
else:
############### 持仓明细表
#file_holding = '组合证券20190319.xlsx' 
    sheet_1 = pd.read_excel(file_holding,sheet_name=0)
 
########### 权益类多头汇总明细表    
#Get_DF(wb['权益类多头汇总明细表'],'A1','U')
    sheet_2 = pd.read_excel(file_holding,sheet_name=1,usecols='A:N')


sheet_1.loc[sheet_1.组合名称 != '沪港通组合','证券代码'] = sheet_1.loc[sheet_1.组合名称 != '沪港通组合','证券代码'].apply(lambda x: str(x).zfill(6))
sheet_1.loc[sheet_1.组合名称 == '沪港通组合','证券代码'] = sheet_1.loc[sheet_1.组合名称 == '沪港通组合','证券代码'].apply(lambda x: str(x).zfill(5))
#Get_DF(wb['持仓明细表'],'A1','U')
Clear_Range(wb['持仓明细表'],'A2:U{}'.format(wb['持仓明细表'].max_row))
Write_DF(wb['持仓明细表'],'A2',sheet_1)


Clear_Range(wb['权益类多头汇总明细表'],'A2:U{}'.format(wb['权益类多头汇总明细表'].max_row))
Write_DF(wb['权益类多头汇总明细表'],'A2',sheet_2)
   

################## 汇总 wb.sheetnames

ws = wb['汇总']

def Get_Prof_Futures(df,Asset_No,T_Day=True):
    col_names = ['账户权益','账户出入金']
    if Asset_No == '量化':
        M_F = df.dropna().query("资产单元名称.str.contains({!r})".format(Asset_No))[col_names]
    else:
        M_F = df.dropna().query("资产单元名称 == {!r}".format(Asset_No))[col_names]   
    
    if T_Day:
        return sum(M_F.iloc[:,0] - M_F.iloc[:,1])
    else:
        return sum(M_F.iloc[:,0].values)
            

ws_2 = wb['期货']

#file_analys = '期货保证金20190319.xls'

futures_analys = pd.read_excel(file_analys)

L_M_F = Get_DF(ws_2,'W1','AI')

ws['L9'].value += Get_Prof_Futures(futures_analys,'量化') - Get_Prof_Futures(L_M_F,'量化',False)
ws['L18'].value += Get_Prof_Futures(futures_analys,'权益类5号场内资产单元') - Get_Prof_Futures(L_M_F,'权益类5号场内资产单元',False)
ws['L19'].value += Get_Prof_Futures(futures_analys,'权益类6号场内资产单元') - Get_Prof_Futures(L_M_F,'权益类6号场内资产单元',False)
ws['L3'].value += Get_Prof_Futures(futures_analys,'权益类场内01资产单元') - Get_Prof_Futures(L_M_F,'权益类场内01资产单元',False)



######################### 期货
# wb_data['期货'].max_column
L_F_H = Get_DF(ws_2,'A1','U')
Clear_Range(wb['期货'],'A2:U{}'.format(wb['期货'].max_row))

#file_futures = '期货持仓查询20190319.xls'

futures_holds = pd.read_excel(file_futures)
if KEEP:
    Write_DF(wb['期货'],'A1',futures_holds[L_F_H.columns],header=True)
else:
    Write_DF(wb['期货'],'A1',futures_holds,header=True)

Clear_Range(wb['期货'],'W2:AI{}'.format(wb['期货'].max_row))
#file_analys = '期货保证金20190319.xls'
#futures_analys = pd.read_excel(file_analys)
if KEEP:
    Write_DF(wb['期货'],'W1',futures_analys[L_M_F.columns],header=True)
else:
    Write_DF(wb['期货'],'W1',futures_analys,header=True)
#Write_DF(wb['期货'],'W1',futures_analys[L_M_F.columns],header=True)
#Get_DF(wb['期货'],'W1','AI')

######################## 资产明细
ws = wb['资产明细']
ws_data = wb_data['资产明细']

for row in ws['H2:I12']:
    for col in row:
        if col.column == 'H':
            col.value = ws_data[col.coordinate].offset(0,-4).value
        else:
            #print(col.value,ws_data[col.coordinate].offset(0,-4)].value)
            col.value = ws_data[col.coordinate].offset(0,-3).value


############# 风险提示

ws= wb['风险提示']
ws_data = wb_data['风险提示']

ws['E14'].value = ws_data['B14'].value
ws['C16'].value = futures_analys['风险比例1(%)'].values[-1] / 100
ws['C27'].value = futures_analys.loc[futures_analys.基金名称 == '自营权益类6号','风险比例1(%)'].values[0] / 100
ws['E27'].value = futures_analys.loc[futures_analys.基金名称 == '自营权益类5号','风险比例1(%)'].values[0] / 100
ws['G27'].value = sum(futures_analys.loc[futures_analys.基金名称 == '自营量化1号','风险比例1(%)'].values) / 100

for col in ws['F19:F21']:
    for row in col:
        #print(row.value,ws_data[row.coordinate].offset(0,-2).value)
        row.value = ws_data[row.coordinate].offset(0,-2).value

#############

##### 可用资金
sel_cols= ['组合名称','证券代码','证券名称','当日买金额','当日卖金额', '当日买费用', '当日卖费用','净买金额','费用合计','证券类别']

#comb_file = '新综合信息查询_组合证券20190319.xls'
#tmp = pd.read_excel(comb_file)
pct = tmp[['证券代码','当日涨跌幅(%)']].dropna().drop_duplicates()
Trades = tmp[sel_cols].query("净买金额 != 0 and 证券类别 in ['股票','开放式基金']")

# 扣除未上市新股市值
mkt_df = tmp[['证券代码','网上新股待上市数量','网下新股待上市数量','市值']].dropna().query('网上新股待上市数量 > 0 or 网下新股待上市数量 > 0')
sh_minus_mkt = mkt_df.query('证券代码.str.startswith("6")')['市值'].sum()
sz_minus_mkt = mkt_df['市值'].sum() - sh_minus_mkt


def Get_Net_Amt(df,sql_str,buy_dir=True):
    a, b, c, d = df.query(sql_str)[['当日卖费用','当日卖金额','当日买费用','当日买金额']].sum().values
    if buy_dir:
        return c + d
    else:
        return b - a

bonus = tmp[['组合名称','证券代码','当日红利']].dropna().query('当日红利 > 0')
sql_seq = iter([sql_acct_0,sql_acct_3,sql_acct_2,sql_acct_1])
for row in wb['可用资金']['D65:E68']:
    sql_curr = next(sql_seq)
    for col in row:
        if col.column == 'D':
            col.value += Get_Net_Amt(Trades,sql_curr,False) + sum(bonus.query(sql_curr)['当日红利'])
        else:
            col.value += Get_Net_Amt(Trades,sql_curr,True)
        #col.value += Get_Net_Amt(Trades,sql_curr,False)
        #print(sql_curr)


'''
#a, b = Trades.query(sql_acct_1)[['当日卖费用','当日卖金额']].sum().values
#b - a
wb['可用资金']['D68'].value += Get_Net_Amt(Trades,sql_acct_1,False)

#a, b = Trades.query(sql_acct_1)[['当日买费用','当日买金额']].sum().values
## a + b
wb['可用资金']['E68'].value += Get_Net_Amt(Trades,sql_acct_1,True)

#a, b = Trades.query(sql_acct_2)[['当日卖费用','当日卖金额']].sum().values
wb['可用资金']['D67'].value += Get_Net_Amt(Trades,sql_acct_2,False)

#a, b = Trades.query(sql_acct_2)[['当日买费用','当日买金额']].sum().values
wb['可用资金']['E67'].value += Get_Net_Amt(Trades,sql_acct_2,True)


#a, b = Trades.query(sql_acct_3)[['当日卖费用','当日卖金额']].sum().values
wb['可用资金']['D66'].value += Get_Net_Amt(Trades,sql_acct_3,False)

#a, b = Trades.query(sql_acct_3)[['当日买费用','当日买金额']].sum().values
wb['可用资金']['E66'].value += Get_Net_Amt(Trades,sql_acct_3,True)


#a, b = Trades.query(sql_acct_0)[['当日卖费用','当日卖金额']].sum().values
wb['可用资金']['D65'].value += Get_Net_Amt(Trades,sql_acct_0,False)

#a, b = Trades.query(sql_acct_0)[['当日买费用','当日买金额']].sum().values
wb['可用资金']['E65'].value += Get_Net_Amt(Trades,sql_acct_0,True)
'''  

######### 单账户1


col_names = ['组合名称','证券名称', '证券代码', '成本价', '成本', '数量', '收盘价', '市值', '浮动盈亏','累积投资收益'] + list(out_cols[-1:])
Acct_1 = sheet_1.query(sql_acct_1).sort_values(['组合名称','市值'],ascending=False)[col_names]

#wb_data.sheetnames

Mins_Profits = Get_DF(wb_data['打新市值统计'],'M1','R').drop(columns=['证券名称'])
Mins_Profits.loc[:,'证券代码'] = Mins_Profits.loc[:,'证券代码'].astype('str')
tmp = pd.merge(Acct_1,Mins_Profits,on=['组合名称','证券代码'],how='left').fillna(0)
tmp['累积投资收益']  -= tmp.iloc[:,-3]
tmp[col_names[-1]] -= tmp.iloc[:,-1]
tmp = tmp[col_names]

#Get_DF(wb['单账户1'],'A2','K')
#### Sort
#tmp.loc[:,'证券代码'] = tmp.loc[:,'证券代码'].apply(lambda x: str(x).zfill(6))
Sort_DF = Get_DF(wb['单账户1'],'A2','C').drop(columns=['证券名称'])
Sort_DF.columns = ['组合名称','证券代码']
#sort_loc = [(x,z) for x, y, z in Sort_DF.values.tolist()]
Sort_DF.loc[:,'证券代码'] = Sort_DF.loc[:,'证券代码'].astype('str')
tmp = pd.merge(Sort_DF,tmp,how='right',on=['组合名称','证券代码'])
tmp = tmp[col_names]

n_row = wb['单账户1'].max_row
Clear_Range(wb['单账户1'],'A3:K{}'.format(n_row))
Write_DF(wb['单账户1'],'A3',tmp)
#Get_DF(wb['单账户1'],'A3','K')

Clear_Range(wb['单账户1'],'N3:W{}'.format(n_row))
sub_cols = ['证券代码','证券名称','成本','市值'] + col_names[-1:]

Agg_tmp = tmp[sub_cols].groupby(['证券代码','证券名称'])[sub_cols[2:]].sum().reset_index()
Agg_tmp.insert(4,'浮动盈亏',Agg_tmp['市值']-Agg_tmp['成本'])
Agg_tmp.sort_values(['市值'],ascending=False,inplace=True)
Agg_tmp = pd.merge(Agg_tmp.query('市值 > 0'),pct,on='证券代码',how='left')
Agg_tmp['收益率'] = Agg_tmp[col_names[-1]] / Agg_tmp['成本']
Agg_tmp = pd.merge(Agg_tmp,INDUSTRY,on='证券代码',how='left')
Agg_tmp['浮动盈亏率'] = Agg_tmp['浮动盈亏'] / Agg_tmp['成本']
Write_DF(wb['单账户1'],'N3',Agg_tmp)

#Get_DF(wb['单账户1'],'N2','V')

######### 单账户2

Acct_2 = sheet_1.query(sql_acct_2).sort_values(['组合名称','市值'],ascending=False)[col_names]

#wb_data.sheetnames

#Mins_Profits = Get_DF(wb_data['打新市值统计'],'M1','Q')
tmp = pd.merge(Acct_2,Mins_Profits,on=['组合名称','证券代码'],how='left').fillna(0)
tmp['累积投资收益']  -= tmp.iloc[:,-3]
tmp[col_names[-1]] -= tmp.iloc[:,-1]
tmp = tmp[col_names]

#Get_DF(wb['单账户2'],'A2','K')
#### Sort
#tmp.loc[:,'证券代码'] = tmp.loc[:,'证券代码'].apply(lambda x: str(x).zfill(6))
Sort_DF = Get_DF(wb['单账户2'],'A2','C').drop(columns=['证券名称'])
Sort_DF.columns = ['组合名称','证券代码']
#sort_loc = [(x,z) for x, y, z in Sort_DF.values.tolist()]
Sort_DF.loc[:,'证券代码'] = Sort_DF.loc[:,'证券代码'].astype('str')
tmp = pd.merge(Sort_DF,tmp,how='right',on=['组合名称','证券代码'])
tmp = tmp[col_names]

n_row = wb['单账户2'].max_row
Clear_Range(wb['单账户2'],'A3:K{}'.format(n_row))
Write_DF(wb['单账户2'],'A3',tmp)
#Get_DF(wb['单账户1'],'A3','K')

Clear_Range(wb['单账户2'],'N3:W{}'.format(n_row))
#sub_cols = ['证券代码','证券名称','成本','市值','总盈亏']

Agg_tmp = tmp[sub_cols].groupby(['证券代码','证券名称'])[sub_cols[2:]].sum().reset_index()
Agg_tmp.insert(4,'浮动盈亏',Agg_tmp['市值']-Agg_tmp['成本'])
Agg_tmp.sort_values(['市值'],ascending=False,inplace=True)
Agg_tmp = pd.merge(Agg_tmp.query('市值 > 0'),pct,on='证券代码',how='left')
Agg_tmp['收益率'] = Agg_tmp[col_names[-1]] / Agg_tmp['成本']
Agg_tmp = pd.merge(Agg_tmp,INDUSTRY,on='证券代码',how='left')
Agg_tmp['浮动盈亏率'] = Agg_tmp['浮动盈亏'] / Agg_tmp['成本']
Write_DF(wb['单账户2'],'N3',Agg_tmp)

#Get_DF(wb['单账户2'],'N2','V')

######### 大账户

Acct_0 = sheet_1.query(sql_acct_0).sort_values(['组合名称','市值'],ascending=False)[col_names]

#wb_data.sheetnames

#Mins_Profits = Get_DF(wb_data['打新市值统计'],'M1:Q43')
tmp = pd.merge(Acct_0,Mins_Profits,on=['组合名称','证券代码'],how='left').fillna(0)
tmp['累积投资收益']  -= tmp.iloc[:,-3]
tmp[col_names[-1]] -= tmp.iloc[:,-1]
tmp = tmp[col_names]
#tmp.loc[:,'证券代码'] = tmp.loc[:,'证券代码'].apply(lambda x: str(x).zfill(6))
#### Sort
#tmp.loc[:,'证券代码'] = tmp.loc[:,'证券代码'].apply(lambda x: str(x).zfill(6))
Sort_DF = Get_DF(wb['大账户'],'A2','C').drop(columns=['证券名称'])
Sort_DF.columns = ['组合名称','证券代码']
#sort_loc = [(x,z) for x, y, z in Sort_DF.values.tolist()]
Sort_DF.loc[:,'证券代码'] = Sort_DF.loc[:,'证券代码'].astype('str')
tmp = pd.merge(Sort_DF,tmp,how='right',on=['组合名称','证券代码'])
tmp = tmp[col_names]

n_row = wb['大账户'].max_row
Clear_Range(wb['大账户'],'A3:K{}'.format(n_row))
Write_DF(wb['大账户'],'A3',tmp)
#Get_DF(wb['大账户'],'A2','K')

Clear_Range(wb['大账户'],'O3:P{}'.format(tmp.shape[0]+2))
val = pd.merge(tmp[['证券代码']],pct,on='证券代码',how='left')
val = pd.merge(val,INDUSTRY,on='证券代码',how='left').drop(columns=['证券代码'])

Write_DF(wb['大账户'],'O3',val)
#### 打新市值统计 Get_DF(wb['大账户'],'O2','P').iloc[:,-1]
### sh
sh_mkt_val = sheet_1.query("组合名称 != '沪港通组合' and 证券代码.str.startswith({!r})".format('6'))['市值'].sum() - sh_minus_mkt
wb['打新市值统计']['J2'].value = sh_mkt_val

### SZ
sz_mkt_val = sheet_1.query("组合名称 != '沪港通组合' and 证券代码.str.startswith({!r})".format('0'))['市值'].sum() + sheet_1.query("组合名称 != '沪港通组合' and 证券代码.str.startswith({!r})".format('3'))['市值'].sum() -sz_minus_mkt
wb['打新市值统计']['J3'].value = sz_mkt_val

nrow, ncol = Get_DF(wb['打新市值统计'],'B1','D').dropna().shape
wb['打新市值统计']['C2'].offset(nrow,0).value = sh_mkt_val
wb['打新市值统计']['D2'].offset(nrow,0).value = sz_mkt_val

#####

date_str_2 = str(datetime.today().year)+'-'+str(datetime.today().month)+'-'+str(datetime.today().day)
file_des = '持仓统计' + date_str + '.xlsx'
wb.save(file_des)

if FLAG:
    flie_lst = (rd_file,file_analys,file_futures,comb_file)
else:
    flie_lst = (rd_file,file_holding,file_analys,file_futures,comb_file)

for file in flie_lst:
    system('move ' + file + ' ' + PATH)

def Sleep_Secs(secs_n=5):
    time_0 = datetime.now()
    while True:
        time_1 = datetime.now()
        if (time_1 - time_0).total_seconds() > secs_n:
            break
        else:
            pass
#Sleep_Secs(5)
system('start OPEN_SAVE.vbs')
#input('Pause:')
#wb = load_workbookfile_des
print('正在填充数据' + '.' * 60)
Sleep_Secs(SECS)
################################ Fill values
        
wb = load_workbook(file_des)
wb_data = load_workbook(file_des,data_only=True)

nrow = Get_DF(wb['汇总'],'U1','V').shape[0]
wb['汇总']['U2'].offset(nrow,0).value = date_str_2
wb['汇总']['V2'].offset(nrow,0).value = wb_data['汇总']['P20'].value

nrow = Get_DF(wb['单账户1'],'AJ1','AJ').shape[0]
wb['单账户1']['AI2'].offset(nrow,0).value = date_str_2
wb['单账户1']['AJ2'].offset(nrow,0).value = wb_data['单账户1']['AG2'].value

nrow = Get_DF(wb['单账户2'],'AJ2','AJ').shape[0]
wb['单账户2']['AJ3'].offset(nrow,0).value = date_str_2
wb['单账户2']['AK3'].offset(nrow,0).value = wb_data['单账户2']['AG2'].value

nrow = Get_DF(wb['单账户3'],'A1','G').shape[0]

wb['单账户3']['A2'].offset(nrow,0).value = date_str_2
wb['单账户3']['A2'].offset(nrow,1).value = '=B{0}+C{0}'.format(nrow+1)
wb['单账户3']['A2'].offset(nrow,2).value = 0.00
wb['单账户3']['A2'].offset(nrow,3).value = '=SUM(C$2:C{0})'.format(nrow+2)
wb['单账户3']['A2'].offset(nrow,4).value = '=E{0}+C{1}/F{0}'.format(nrow+1,nrow+2)
wb['单账户3']['A2'].offset(nrow,5).value = '=($I$1+G{0}+D{0})/E{0}'.format(nrow+2)
wb['单账户3']['A2'].offset(nrow,6).value = wb_data['汇总']['P9'].value
wb_data.close()
wb.save(file_des)

system('start OPEN_SAVE.vbs')
print('正在生成单账户表' + '.' * 56)
Sleep_Secs(SECS)
wb_data = load_workbook(file_des,data_only=True)

workbook = load_workbook('Acct1.xlsx')
new_sheet = workbook.active
new_sheet.title = '单账户1'
default_sheet = wb_data['单账户1']

for row in default_sheet.rows:
    for cell in row:
        new_cell = new_sheet.cell(row=cell.row, column=cell.col_idx,value= cell.value)
        if cell.has_style:
            new_cell.font = copy(cell.font)
            new_cell.border = copy(cell.border)
            new_cell.fill = copy(cell.fill)
            new_cell.number_format = copy(cell.number_format)
            new_cell.protection = copy(cell.protection)
            new_cell.alignment = copy(cell.alignment)
workbook.save('单账户1-' + date_str + '.xlsx')

workbook = load_workbook('Acct2.xlsx')
new_sheet = workbook.active
new_sheet.title = '单账户2'
default_sheet = wb_data['单账户2']

for row in default_sheet.rows:
    for cell in row:
        new_cell = new_sheet.cell(row=cell.row, column=cell.col_idx,value= cell.value)
        if cell.has_style:
            new_cell.font = copy(cell.font)
            new_cell.border = copy(cell.border)
            new_cell.fill = copy(cell.fill)
            new_cell.number_format = copy(cell.number_format)
            new_cell.protection = copy(cell.protection)
            new_cell.alignment = copy(cell.alignment)
#default_sheet.column_dimensions['A'].width
#new_sheet.column_dimensions['A'].width
workbook.save('单账户2-' + date_str + '.xlsx')
print('已完成.\n' + '.' * 72)
system('start '+ file_des)
input('Enter:')

